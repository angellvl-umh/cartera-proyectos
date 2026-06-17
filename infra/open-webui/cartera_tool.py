"""
title: Cartera de Proyectos TIC
description: Herramientas para consultar y gestionar la cartera de proyectos TIC universitaria. Permite al agente IA consultar tareas, proyectos, capacidad de equipos, registrar avance semanal y consultar/exportar el informe semanal de cartera, realizar acciones como cambiar estados o crear tareas, generar gráficos visuales (tarta o barras) y exportar listados a Excel.
author: Cartera TIC
version: 1.3.0
required_open_webui_version: 0.5.0
requirements: matplotlib, openpyxl
"""

import io
import json
import requests

import matplotlib
matplotlib.use("Agg")  # backend sin display, obligatorio en Docker
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from typing import Optional


class Tools:
    class Valves(BaseModel):
        base_url: str = "http://backend:8080/api/agent"
        api_key: str = "cartera-agent-key"
        external_url: str = "http://localhost:5000"  # URL pública del backend (accesible desde el navegador)

    def __init__(self):
        self.valves = self.Valves()

    def _headers(self, email: str) -> dict:
        return {
            "Authorization": f"Bearer {self.valves.api_key}",
            "X-Open-WebUI-User-Email": email,
            "Content-Type": "application/json",
        }

    def _get(self, path: str, email: str, params: dict = None) -> dict:
        try:
            r = requests.get(
                f"{self.valves.base_url}{path}",
                headers=self._headers(email),
                params=params,
                timeout=10,
            )
            return r.json() if r.ok else {"error": r.status_code, "detail": r.text}
        except Exception as e:
            return {"error": str(e)}

    def _post(self, path: str, email: str, body: dict = None) -> dict:
        try:
            r = requests.post(
                f"{self.valves.base_url}{path}",
                headers=self._headers(email),
                json=body or {},
                timeout=10,
            )
            return r.json() if r.ok else {"error": r.status_code, "detail": r.text}
        except Exception as e:
            return {"error": str(e)}

    def get_my_tasks(self, __user__: dict) -> str:
        """
        Obtener mis tareas pendientes.
        Devuelve las tareas activas, en backlog y completadas del usuario.
        Úsalo cuando el usuario pregunte '¿qué tengo pendiente?', '¿en qué estoy trabajando?', '¿cuáles son mis tareas?'
        """
        result = self._get("/me", __user__.get("email", ""))
        return json.dumps(result, ensure_ascii=False)

    def get_projects(
        self,
        __user__: dict,
        sipt_group: Optional[str] = None,
        status: Optional[str] = None,
    ) -> str:
        """
        Listar proyectos del usuario, con filtros opcionales.
        Devuelve los proyectos del usuario con su estado y progreso de tareas.
        Úsalo cuando el usuario pregunte por sus proyectos o el estado general de la cartera.
        :param sipt_group: Filtrar por grupo SIPT. Valores válidos: WebTransversal, RRHH, Academico, Sede, Observatorio, InvestigacionEconomico
        :param status: Filtrar por estado. Valores válidos: Stopped, PlanningWithClient, WaitingForDevelopers, PlanningSprint, InSprint, DevelopmentOutsideSprint, InTesting, Completed, PostponedByClient
        """
        params: dict = {}
        if sipt_group:
            params["siptGroup"] = sipt_group
        if status:
            params["status"] = status
        result = self._get("/projects", __user__.get("email", ""), params=params or None)
        return json.dumps(result, ensure_ascii=False)

    def get_project_detail(self, project_id: int, __user__: dict) -> str:
        """
        Obtener detalle de un proyecto.
        Devuelve información detallada: estado, sprints activos, tareas pendientes y progreso.
        Úsalo cuando el usuario pregunte '¿cómo va el proyecto X?' o pida info sobre un proyecto concreto.
        :param project_id: ID numérico del proyecto
        """
        result = self._get(f"/projects/{project_id}", __user__.get("email", ""))
        return json.dumps(result, ensure_ascii=False)

    def get_capacity(self, __user__: dict) -> str:
        """
        Consultar carga y disponibilidad de equipos.
        Devuelve la carga de trabajo de todos los equipos y sus miembros.
        Green = disponible (≤3 tareas activas), Yellow = cargado (4-6), Red = saturado (≥7).
        Úsalo cuando el usuario pregunte qué equipo tiene disponibilidad para un nuevo proyecto.
        """
        result = self._get("/capacity", __user__.get("email", ""))
        return json.dumps(result, ensure_ascii=False)

    def search_tasks(self, query: str, __user__: dict) -> str:
        """
        Buscar tarea por descripción usando búsqueda semántica.
        Devuelve las 5 tareas más similares al texto proporcionado.
        Úsalo para identificar una tarea concreta cuando el usuario la mencione de forma natural,
        por ejemplo 'la tarea del proxy SSL' o 'lo del certificado'.
        :param query: Texto a buscar de forma natural
        """
        result = self._get("/tasks/search", __user__.get("email", ""), params={"q": query})
        return json.dumps(result, ensure_ascii=False)

    def update_task_status(self, task_id: int, status: str, __user__: dict) -> str:
        """
        Cambiar el estado de una tarea.
        Estados válidos: Backlog, ToDo, InProgress, Blocked, Done. Done es terminal y no puede revertirse.
        IMPORTANTE: confirma siempre con el usuario antes de ejecutar este cambio.
        Úsalo cuando el usuario diga que terminó una tarea, que está bloqueado o que empieza a trabajar en algo.
        :param task_id: ID numérico de la tarea
        :param status: Nuevo estado — uno de: Backlog, ToDo, InProgress, Blocked, Done
        """
        result = self._post(
            f"/tasks/{task_id}/status",
            __user__.get("email", ""),
            {"status": status},
        )
        return json.dumps(result, ensure_ascii=False)

    def create_task(
        self,
        project_id: int,
        title: str,
        __user__: dict,
        description: Optional[str] = None,
        priority: Optional[str] = "Medium",
        epic_id: Optional[int] = None,
        sprint_id: Optional[int] = None,
        assign_to_self: Optional[bool] = True,
    ) -> str:
        """
        Crear una nueva tarea en un proyecto.
        Úsalo cuando el usuario quiera registrar trabajo pendiente.
        :param project_id: ID numérico del proyecto
        :param title: Título de la tarea
        :param description: Descripción opcional
        :param priority: Prioridad — Low, Medium, High o Critical. Por defecto Medium.
        :param epic_id: ID de la épica (opcional)
        :param sprint_id: ID del sprint al que asignar la tarea (opcional)
        :param assign_to_self: Si true (por defecto), asigna la tarea al usuario que la crea
        """
        body = {
            "projectId": project_id,
            "title": title,
            "description": description,
            "priority": priority,
            "epicId": epic_id,
            "sprintId": sprint_id,
            "assignToSelf": assign_to_self,
        }
        result = self._post("/tasks", __user__.get("email", ""), body)
        return json.dumps(result, ensure_ascii=False)

    def add_task_comment(self, task_id: int, text: str, __user__: dict) -> str:
        """
        Añadir un comentario o nota de seguimiento a una tarea existente.
        El comentario queda registrado con el autor y la fecha actual.
        Úsalo cuando el usuario quiera documentar avances, bloqueos o novedades sobre una tarea.
        :param task_id: ID numérico de la tarea
        :param text: Texto del comentario
        """
        result = self._post(
            f"/tasks/{task_id}/comment",
            __user__.get("email", ""),
            {"text": text},
        )
        return json.dumps(result, ensure_ascii=False)

    def add_project_note(self, project_id: int, text: str, __user__: dict) -> str:
        """
        Añadir una nota de seguimiento a un proyecto.
        La nota queda registrada con el autor y la fecha actual.
        Úsalo cuando el usuario quiera documentar decisiones, hitos, bloqueos o novedades sobre un proyecto concreto.
        :param project_id: ID numérico del proyecto
        :param text: Texto de la nota de seguimiento
        """
        result = self._post(
            f"/projects/{project_id}/notes",
            __user__.get("email", ""),
            {"text": text},
        )
        return json.dumps(result, ensure_ascii=False)

    def add_weekly_update(
        self,
        project_id: int,
        summary: str,
        health_status: str,
        __user__: dict,
    ) -> str:
        """
        Registrar el avance semanal del usuario en un proyecto.
        Si ya existe un registro de esta persona en este proyecto para la semana ISO actual, lo actualiza en vez de duplicarlo.
        Úsalo cuando el usuario quiera reportar cómo va un proyecto esta semana, p.ej. 'añade mi avance semanal al proyecto X diciendo que vamos bien'.
        :param project_id: ID numérico del proyecto
        :param summary: Resumen del avance de esta semana (máx. 1000 caracteres)
        :param health_status: Estado de salud del proyecto — OnTrack (en curso), AtRisk (en riesgo) o Blocked (bloqueado)
        """
        result = self._post(
            f"/projects/{project_id}/weekly-updates",
            __user__.get("email", ""),
            {"summary": summary, "healthStatus": health_status},
        )
        return json.dumps(result, ensure_ascii=False)

    def get_weekly_portfolio_report(
        self,
        __user__: dict,
        year: Optional[int] = None,
        team_id: Optional[int] = None,
        sipt_group: Optional[str] = None,
    ) -> str:
        """
        Obtener el informe semanal de seguimiento de la cartera.
        Devuelve los proyectos en curso separados en 'en riesgo' (sin actualización esta semana o con estado AtRisk/Blocked) y el resto, cada uno con su último avance registrado.
        Úsalo cuando el usuario pida un resumen del estado de la cartera esta semana o pregunte qué proyectos están en riesgo.
        :param year: Filtrar por año de cartera (opcional)
        :param team_id: Filtrar por ID de equipo (opcional)
        :param sipt_group: Filtrar por grupo SIPT. Valores válidos: WebTransversal, RRHH, Academico, Sede, Observatorio, InvestigacionEconomico
        """
        params: dict = {}
        if year is not None:
            params["year"] = year
        if team_id is not None:
            params["teamId"] = team_id
        if sipt_group:
            params["siptGroup"] = sipt_group
        result = self._get("/weekly-portfolio-report", __user__.get("email", ""), params=params or None)
        return json.dumps(result, ensure_ascii=False)

    def export_weekly_portfolio_report_excel(
        self,
        __user__: dict,
        year: Optional[int] = None,
        team_id: Optional[int] = None,
        sipt_group: Optional[str] = None,
    ) -> str:
        """
        Exportar el informe semanal de seguimiento de la cartera a un fichero Excel descargable.
        Aplica los mismos filtros opcionales que get_weekly_portfolio_report. Los proyectos en riesgo aparecen primero, marcados en una columna aparte.
        Úsalo cuando el usuario pida 'exportar el informe semanal a Excel' o 'descargar el seguimiento de la cartera'.
        :param year: Filtrar por año de cartera (opcional)
        :param team_id: Filtrar por ID de equipo (opcional)
        :param sipt_group: Filtrar por grupo SIPT. Valores válidos: WebTransversal, RRHH, Academico, Sede, Observatorio, InvestigacionEconomico
        """
        params: dict = {}
        if year is not None:
            params["year"] = year
        if team_id is not None:
            params["teamId"] = team_id
        if sipt_group:
            params["siptGroup"] = sipt_group
        data = self._get("/weekly-portfolio-report", __user__.get("email", ""), params=params or None)
        if isinstance(data, dict) and "error" in data:
            return json.dumps(data, ensure_ascii=False)

        at_risk = data.get("atRiskProjects", [])
        other = data.get("otherProjects", [])
        if not at_risk and not other:
            return "No hay proyectos que coincidan con el filtro indicado."

        headers = ["En riesgo", "Proyecto", "Equipo principal", "Estado", "Estado de salud",
                   "Última actualización", "Autor", "Fecha", "Resumen"]
        wb = Workbook()
        ws = wb.active
        ws.title = "Informe semanal"
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for p in at_risk + other:
            ws.append([
                "Sí" if p.get("isAtRisk") else "No",
                p.get("title"), p.get("primaryTeamName"), p.get("status"),
                p.get("latestHealthStatus"),
                "Sí" if p.get("hasUpdateThisWeek") else "No",
                p.get("latestAuthorName"), p.get("latestUpdateDate"),
                p.get("latestSummary"),
            ])
        for i, header in enumerate(headers, start=1):
            width = max(len(header) + 2, 12)
            ws.column_dimensions[get_column_letter(i)].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return self._upload_export(buf.getvalue(), "informe-semanal-cartera.xlsx")

    def reindex(self, __user__: dict) -> str:
        """
        Regenerar el índice de embeddings para búsqueda semántica.
        Genera o actualiza los vectores de todas las tareas. Ejecutar después de crear o modificar tareas masivamente.
        Puede tardar varios segundos según el número de tareas.
        """
        result = self._post("/reindex", __user__.get("email", ""))
        return json.dumps(result, ensure_ascii=False)

    # ── Exports (Excel) ──────────────────────────────────────────────────────

    def _upload_export(self, data: bytes, file_name: str) -> str:
        r = requests.post(
            f"{self.valves.base_url}/exports",
            headers={"Authorization": f"Bearer {self.valves.api_key}",
                     "Content-Type": "application/octet-stream"},
            params={"fileName": file_name},
            data=data,
            timeout=15,
        )
        if not r.ok:
            return f"Error guardando el Excel: {r.status_code}"
        export_id = r.json()["id"]
        url = f"{self.valves.external_url}/api/agent/exports/{export_id}"
        return f"[{file_name}]({url})"

    def export_projects_excel(
        self,
        __user__: dict,
        sipt_group: Optional[str] = None,
        status: Optional[str] = None,
    ) -> str:
        """
        Exportar el listado de proyectos a un fichero Excel descargable.
        Aplica los mismos filtros opcionales que get_projects. No tiene paginación: incluye todos los proyectos que cumplan el filtro.
        Úsalo cuando el usuario pida 'exportar a Excel', 'descargar el listado' o 'pasar esto a una hoja de cálculo'.
        :param sipt_group: Filtrar por grupo SIPT. Valores válidos: WebTransversal, RRHH, Academico, Sede, Observatorio, InvestigacionEconomico
        :param status: Filtrar por estado. Valores válidos: Stopped, PlanningWithClient, WaitingForDevelopers, PlanningSprint, InSprint, DevelopmentOutsideSprint, InTesting, Completed, PostponedByClient
        """
        params: dict = {}
        if sipt_group:
            params["siptGroup"] = sipt_group
        if status:
            params["status"] = status
        data = self._get("/projects", __user__.get("email", ""), params=params or None)
        if isinstance(data, dict) and "error" in data:
            return json.dumps(data, ensure_ascii=False)
        if not data:
            return "No hay proyectos que coincidan con el filtro indicado."

        headers = ["ID", "Título", "Estado", "Unidad solicitante", "Equipo principal",
                   "Tareas totales", "Tareas hechas", "Sprints activos"]
        wb = Workbook()
        ws = wb.active
        ws.title = "Proyectos"
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for p in data:
            ws.append([
                p.get("id"), p.get("title"), p.get("status"),
                p.get("requestingUnit"), p.get("primaryTeamName"),
                p.get("totalTasks"), p.get("doneTasks"), p.get("activeSprints"),
            ])
        for i, header in enumerate(headers, start=1):
            width = max(len(header) + 2, 12)
            ws.column_dimensions[get_column_letter(i)].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return self._upload_export(buf.getvalue(), "proyectos.xlsx")

    # ── Charts ────────────────────────────────────────────────────────────────

    def _plot_categories(self, ax_or_fig, labels, values, colors, chart_type: str, title: str):
        """Dibuja barras horizontales o tarta, según chart_type ('bar' o 'pie')."""
        if chart_type == "pie":
            fig, ax = plt.subplots(figsize=(5.5, 4.5))
            total = sum(values)
            wedges, _, autotexts = ax.pie(
                values, labels=None, colors=colors,
                autopct=lambda p: f"{p:.0f}%\n({int(round(p * total / 100))})",
                pctdistance=0.72, startangle=90,
                wedgeprops={"edgecolor": "white", "linewidth": 2},
            )
            for at in autotexts:
                at.set_fontsize(8)
            ax.legend(wedges, labels, loc="lower center", bbox_to_anchor=(0.5, -0.1),
                      ncol=2, fontsize=8, frameon=False)
            ax.set_title(title, fontweight="bold", pad=14)
            fig.tight_layout()
            return fig

        fig, ax = plt.subplots(figsize=(7, max(3, len(labels) * 0.45)))
        bars = ax.barh(labels, values, color=colors, edgecolor="white", height=0.6)
        ax.bar_label(bars, padding=4, fontsize=9)
        ax.set_title(title, fontweight="bold", pad=12)
        ax.invert_yaxis()
        ax.set_xlim(0, max(values or [1]) + 2)
        ax.spines[["top", "right"]].set_visible(False)
        fig.tight_layout()
        return fig

    def _fig_to_md(self, fig) -> str:
        buf = io.BytesIO()
        fig.savefig(buf, format="png", bbox_inches="tight", dpi=96)
        plt.close(fig)
        buf.seek(0)
        # Subir el PNG al backend y devolver solo una URL corta (~80 chars)
        # para no inflar el contexto del LLM con miles de tokens de base64
        r = requests.post(
            f"{self.valves.base_url}/charts",
            headers={"Authorization": f"Bearer {self.valves.api_key}",
                     "Content-Type": "application/octet-stream"},
            data=buf.getvalue(),
            timeout=15,
        )
        if not r.ok:
            return f"Error guardando gráfico: {r.status_code}"
        chart_id = r.json()["id"]
        url = f"{self.valves.external_url}/api/agent/charts/{chart_id}"
        return f"![gráfico]({url})"

    def chart_team_capacity(self, __user__: dict) -> str:
        """
        Generar gráfico de carga de trabajo por equipo.
        Muestra las tareas activas de cada miembro con código de color:
        verde (≤3 tareas, disponible), amarillo (4-6, cargado), rojo (≥7, saturado).
        Úsalo cuando el usuario pida ver la capacidad de los equipos de forma visual.
        """
        data = self._get("/capacity", __user__.get("email", ""))
        if "error" in data:
            return json.dumps(data, ensure_ascii=False)

        COLOR_MAP = {"Green": "#4caf50", "Yellow": "#ff9800", "Red": "#f44336"}
        rows = []

        for team in data.get("teams", []):
            for m in team.get("members", []):
                rows.append({
                    "label": f"{m['name']} ({team['teamName']})",
                    "value": m["activeTasks"],
                    "color": COLOR_MAP.get(m["loadLevel"], "#9e9e9e"),
                })

        if not rows:
            return "No hay datos de capacidad disponibles."

        # Ordenar por carga descendente y limitar a 15 para no exceder el contexto
        rows.sort(key=lambda r: r["value"], reverse=True)
        rows = rows[:15]
        names  = [r["label"] for r in rows]
        values = [r["value"] for r in rows]
        colors = [r["color"] for r in rows]

        fig, ax = plt.subplots(figsize=(7, max(3, len(names) * 0.4)))
        bars = ax.barh(names, values, color=colors, edgecolor="white", height=0.6)
        ax.bar_label(bars, padding=4, fontsize=9)
        ax.set_xlabel("Tareas activas")
        ax.set_title("Carga de trabajo por persona", fontweight="bold", pad=12)
        ax.invert_yaxis()
        ax.set_xlim(0, max(values or [1]) + 2)
        ax.spines[["top", "right"]].set_visible(False)

        legend = [
            mpatches.Patch(color="#4caf50", label="Disponible (≤3)"),
            mpatches.Patch(color="#ff9800", label="Cargado (4-6)"),
            mpatches.Patch(color="#f44336", label="Saturado (≥7)"),
        ]
        ax.legend(handles=legend, loc="lower right", fontsize=8)
        fig.tight_layout()
        return self._fig_to_md(fig)

    def chart_project_progress(self, __user__: dict) -> str:
        """
        Generar gráfico de progreso de proyectos.
        Muestra tareas completadas vs pendientes para cada proyecto del usuario.
        Úsalo cuando el usuario quiera ver visualmente el avance de los proyectos.
        """
        data = self._get("/projects", __user__.get("email", ""))
        if "error" in data or not isinstance(data, list) or not data:
            return "No hay proyectos disponibles para mostrar."

        titles = [p["title"][:28] + "…" if len(p["title"]) > 28 else p["title"] for p in data]
        done    = [p["doneTasks"] for p in data]
        pending = [p["totalTasks"] - p["doneTasks"] for p in data]

        x = range(len(titles))
        fig, ax = plt.subplots(figsize=(max(6, len(titles) * 1.1), 4))
        w = 0.38
        b1 = ax.bar([i - w / 2 for i in x], done,    width=w, label="Completadas", color="#4caf50", edgecolor="white")
        b2 = ax.bar([i + w / 2 for i in x], pending, width=w, label="Pendientes",  color="#90a4ae", edgecolor="white")
        ax.bar_label(b1, padding=3, fontsize=8)
        ax.bar_label(b2, padding=3, fontsize=8)
        ax.set_xticks(list(x))
        ax.set_xticklabels(titles, rotation=22, ha="right", fontsize=9)
        ax.set_ylabel("Tareas")
        ax.set_title("Progreso de proyectos", fontweight="bold", pad=12)
        ax.legend(fontsize=9)
        ax.spines[["top", "right"]].set_visible(False)
        fig.tight_layout()
        return self._fig_to_md(fig)

    def chart_my_tasks_by_status(self, __user__: dict, chart_type: Optional[str] = "donut") -> str:
        """
        Generar gráfico de distribución de mis tareas por estado.
        Muestra las tareas del usuario agrupadas por estado, en formato donut (por defecto) o barras.
        Úsalo cuando el usuario quiera ver un resumen visual de su carga de trabajo actual.
        :param chart_type: Tipo de gráfico — 'donut' (por defecto) o 'bar'
        """
        data = self._get("/me", __user__.get("email", ""))
        if "error" in data:
            return json.dumps(data, ensure_ascii=False)

        STATUS_COLORS = {
            "InProgress": "#2196f3",
            "ToDo":       "#ff9800",
            "Backlog":    "#9e9e9e",
            "Blocked":    "#f44336",
            "Done":       "#4caf50",
        }
        STATUS_LABELS = {
            "InProgress": "En progreso",
            "ToDo":       "Por hacer",
            "Backlog":    "Backlog",
            "Blocked":    "Bloqueada",
            "Done":       "Hecha",
        }

        counts: dict[str, int] = {}
        for task in data.get("activeTasks", []):
            counts[task["status"]] = counts.get(task["status"], 0) + 1
        for task in data.get("backlogTasks", []):
            counts["Backlog"] = counts.get("Backlog", 0) + 1
        done_count = data.get("doneTasks", 0)
        if done_count:
            counts["Done"] = done_count

        if not counts:
            return "No tienes tareas registradas."

        labels  = [STATUS_LABELS.get(k, k) for k in counts]
        sizes   = list(counts.values())
        colors  = [STATUS_COLORS.get(k, "#bdbdbd") for k in counts]
        total   = sum(sizes)
        title   = f"Mis tareas — {data.get('userName', '')}"

        if chart_type == "bar":
            fig = self._plot_categories(None, labels, sizes, colors, "bar", title)
            return self._fig_to_md(fig)

        fig, ax = plt.subplots(figsize=(5, 4))
        wedges, _, autotexts = ax.pie(
            sizes, labels=None, colors=colors,
            autopct=lambda p: f"{p:.0f}%\n({int(round(p * total / 100))})",
            pctdistance=0.72, startangle=90,
            wedgeprops={"edgecolor": "white", "linewidth": 2},
        )
        for at in autotexts:
            at.set_fontsize(8)

        centre = plt.Circle((0, 0), 0.48, color="white")
        ax.add_patch(centre)
        ax.text(0, 0, f"{total}\ntareas", ha="center", va="center", fontsize=11, fontweight="bold")
        ax.legend(wedges, labels, loc="lower center", bbox_to_anchor=(0.5, -0.08),
                  ncol=3, fontsize=8, frameon=False)
        ax.set_title(title, fontweight="bold", pad=14)
        fig.tight_layout()
        return self._fig_to_md(fig)

    def chart_projects_by_status(
        self,
        __user__: dict,
        chart_type: Optional[str] = "pie",
        sipt_group: Optional[str] = None,
    ) -> str:
        """
        Generar gráfico de proyectos agrupados por estado.
        Muestra cuántos proyectos hay en cada estado de la cartera (o de los proyectos visibles para el usuario), en formato tarta (por defecto) o barras.
        Úsalo cuando el usuario quiera ver de forma visual cómo se reparte la cartera por estado.
        :param chart_type: Tipo de gráfico — 'pie' (por defecto) o 'bar'
        :param sipt_group: Filtrar por grupo SIPT. Valores válidos: WebTransversal, RRHH, Academico, Sede, Observatorio, InvestigacionEconomico
        """
        params = {"siptGroup": sipt_group} if sipt_group else None
        data = self._get("/projects", __user__.get("email", ""), params=params)
        if isinstance(data, dict) and "error" in data:
            return json.dumps(data, ensure_ascii=False)
        if not data:
            return "No hay proyectos disponibles para mostrar."

        STATUS_LABELS = {
            "Stopped": "Parado", "PlanningWithClient": "Planif. cliente",
            "WaitingForDevelopers": "Esperando dev.", "PlanningSprint": "Planif. sprint",
            "InSprint": "En sprint", "DevelopmentOutsideSprint": "Fuera de sprint",
            "InTesting": "En pruebas", "Completed": "Finalizado", "PostponedByClient": "Pospuesto",
        }
        PALETTE = ["#1890ff", "#d4b106", "#13c2c2", "#52c41a", "#2f54eb",
                   "#fa8c16", "#722ed1", "#ff4d4f", "#595959"]

        counts: dict[str, int] = {}
        for p in data:
            counts[p["status"]] = counts.get(p["status"], 0) + 1

        labels = [STATUS_LABELS.get(k, k) for k in counts]
        values = list(counts.values())
        colors = [PALETTE[i % len(PALETTE)] for i in range(len(counts))]

        fig = self._plot_categories(None, labels, values, colors, chart_type or "pie", "Proyectos por estado")
        return self._fig_to_md(fig)

    def chart_projects_by_team(
        self,
        __user__: dict,
        chart_type: Optional[str] = "bar",
        sipt_group: Optional[str] = None,
        status: Optional[str] = None,
    ) -> str:
        """
        Generar gráfico de proyectos agrupados por equipo principal.
        Muestra cuántos proyectos tiene asignado cada equipo de desarrolladores, en formato barras (por defecto) o tarta.
        Úsalo cuando el usuario quiera ver qué equipo lleva más proyectos o cómo se reparte la carga entre equipos.
        :param chart_type: Tipo de gráfico — 'bar' (por defecto) o 'pie'
        :param sipt_group: Filtrar por grupo SIPT. Valores válidos: WebTransversal, RRHH, Academico, Sede, Observatorio, InvestigacionEconomico
        :param status: Filtrar por estado del proyecto
        """
        params: dict = {}
        if sipt_group:
            params["siptGroup"] = sipt_group
        if status:
            params["status"] = status
        data = self._get("/projects", __user__.get("email", ""), params=params or None)
        if isinstance(data, dict) and "error" in data:
            return json.dumps(data, ensure_ascii=False)
        if not data:
            return "No hay proyectos disponibles para mostrar."

        PALETTE = ["#1890ff", "#52c41a", "#fa8c16", "#722ed1", "#13c2c2", "#ff4d4f", "#2f54eb"]
        counts: dict[str, int] = {}
        for p in data:
            team = p.get("primaryTeamName") or "Sin equipo"
            counts[team] = counts.get(team, 0) + 1

        labels = list(counts.keys())
        values = list(counts.values())
        colors = [PALETTE[i % len(PALETTE)] for i in range(len(counts))]

        fig = self._plot_categories(None, labels, values, colors, chart_type or "bar", "Proyectos por equipo")
        return self._fig_to_md(fig)
